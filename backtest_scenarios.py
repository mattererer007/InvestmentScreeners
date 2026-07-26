"""
backtest_scenarios.py — Oversold Entry → Overbought Exit Backtest
==================================================================
Given a starting date, scans EVERY completed quarter from that date
through today. Any stock that triggers the oversold filter + crossover
signal on any quarter gets an entry at that quarter's close. A stock
can trigger multiple times across different quarters.

Returns are measured under 3 exit scenarios:

  1. Hold Through Today — never sell
  2. Sell on Quarterly Overbought Reversal — sell when 2 of 3 quarterly
     indicators first rise ABOVE overbought, then fall back BELOW
  3. Sell on Monthly Overbought Reversal — sell when 2 of 3 monthly
     indicators first rise ABOVE overbought, then fall back BELOW

Overbought thresholds:
    SMI > 40
    Stochastic Slow avg (%K+%D)/2 > 80
    RSI > 70

The exit is a two-path trigger for scenarios 2 and 3:

    Path A — "Fell Below Average - Sold" (optional, --fell-below/--no-fell-below):
        Before ever reaching overbought, if Stochastic %K falls below %D
        AND RSI falls below its signal line for N consecutive bars
        (--fell-below-bars N), sell (momentum lost).

    Path B — "Overbought - Sold":
        If 2 of 3 indicators go ABOVE overbought (stock is hot), stay in.
        Sell at the close of the first bar where fewer than 2 of 3 remain
        above overbought (momentum fading after the run).

    Once a stock enters overbought territory, the fell-below-average check
    is disabled — you only sell on the overbought reversal.

If a stock never triggers either exit after entry, it stays as
"Still Holding" with return measured to today's price.

Output: Excel workbook with 4 tabs:
    - Summary
    - Hold Through Today
    - Sell Quarterly Overbought
    - Sell Monthly Overbought

Usage:
    python backtest_scenarios.py --date "Q1 2015"
    python backtest_scenarios.py --date "Q4 2008"
    python backtest_scenarios.py --date "March 2020" --no-fell-below
    python backtest_scenarios.py --date "2015-06-30" --fell-below-bars 2

Notes:
    - Survivorship bias: only tests stocks still listed today.
    - Each quarter from --date through the last completed quarter is checked.

Dependencies:
    pip install yfinance pandas requests openpyxl
"""

import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
import yfinance as yf
import requests
import io
import time
import random
import sys
import argparse
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

from indicators import resample_ohlcv, compute_all
from constants import (
    RSI_PERIOD, RSI_SIGNAL_PERIOD,
    STOCH_K_PERIOD, STOCH_SLOW_K, STOCH_SLOW_D,
    SMI_PERIOD, SMI_SMOOTH1, SMI_SMOOTH2,
    SMI_THRESHOLD, STOCH_AVG_THRESHOLD, RSI_AVG_THRESHOLD,
    HISTORY_YEARS, MIN_PRICE, MIN_AVG_VOLUME, BATCH_SIZE,
)


# ══════════════════════════════════════════════════════════════
#  OVERBOUGHT THRESHOLDS (for exit)
# ══════════════════════════════════════════════════════════════

SMI_OVERBOUGHT   = 40
STOCH_OVERBOUGHT = 80
RSI_OVERBOUGHT   = 70


# ══════════════════════════════════════════════════════════════
#  DATE PARSING
# ══════════════════════════════════════════════════════════════

QUARTER_MAP = {
    'Q1': '03-31', 'Q2': '06-30', 'Q3': '09-30', 'Q4': '12-31',
}
MONTH_MAP = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12,
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
    'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9,
    'oct': 10, 'nov': 11, 'dec': 12,
}


def parse_date(raw: str) -> datetime:
    raw = raw.strip()
    if raw.lower() == 'today':
        return datetime.now()

    m = re.match(r'^(Q[1-4])\s+(\d{4})$', raw, re.IGNORECASE)
    if m:
        q, year = m.group(1).upper(), int(m.group(2))
        return datetime.strptime(f"{year}-{QUARTER_MAP[q]}", "%Y-%m-%d")

    m = re.match(r'^([A-Za-z]+)\s+(\d{4})$', raw)
    if m:
        month_str, year = m.group(1).lower(), int(m.group(2))
        if month_str in MONTH_MAP:
            month = MONTH_MAP[month_str]
            if month == 12:
                return datetime(year, 12, 31)
            else:
                return datetime(year, month + 1, 1) - timedelta(days=1)

    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue

    print(f"  Could not parse date '{raw}'.")
    print(f"  Examples: Q1 2015, Q4 2008, March 2020, 2015-06-30")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════
#  EXCHANGE PROMPT
# ══════════════════════════════════════════════════════════════

def prompt_exchange() -> list[str]:
    print("  Which exchange(s) do you want to backtest?")
    print("    1) NYSE only")
    print("    2) NASDAQ only")
    print("    3) Both NYSE + NASDAQ")
    print()

    while True:
        choice = input("  Enter 1, 2, or 3: ").strip()
        if choice == "1":
            return ["NYSE"]
        elif choice == "2":
            return ["NASDAQ"]
        elif choice == "3":
            return ["NYSE", "NASDAQ"]
        else:
            print("  Invalid choice. Please enter 1, 2, or 3.")


# ══════════════════════════════════════════════════════════════
#  UNIVERSE
# ══════════════════════════════════════════════════════════════

def get_stock_tickers(exchanges: list[str]) -> list[str]:
    url = "https://www.alphavantage.co/query?function=LISTING_STATUS&apikey=demo"
    label = " + ".join(exchanges)
    print(f"  Fetching {label} stock listing from Alpha Vantage...")
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        df = pd.read_csv(io.StringIO(resp.text))
        stocks = df[df["assetType"] == "Stock"].copy()
        stocks = stocks[stocks["exchange"].isin(exchanges)]
        if "status" in stocks.columns:
            stocks = stocks[stocks["status"] == "Active"]
        tickers = [t for t in stocks["symbol"].tolist()
                   if isinstance(t, str) and len(t) <= 5 and t.isalpha()]
        print(f"  Found {len(tickers)} stocks on {label}")
        return tickers
    except Exception as e:
        print(f"  Alpha Vantage failed: {e}")
        return [
            "AAPL", "MSFT", "GOOGL", "AMZN", "NVDA", "META", "TSLA",
            "JPM", "BAC", "WFC", "JNJ", "UNH", "PFE", "XOM", "CVX",
            "WMT", "PG", "KO", "DIS", "NFLX", "INTC", "AMD", "CRM",
            "BA", "CAT", "GE", "HD", "MCD", "NKE", "V", "MA",
        ]


# ══════════════════════════════════════════════════════════════
#  BATCH DOWNLOAD
# ══════════════════════════════════════════════════════════════

def download_batch(tickers, start, end):
    try:
        data = yf.download(
            tickers=tickers, start=start, end=end,
            interval="1d", group_by="ticker",
            auto_adjust=True, progress=False, threads=True,
        )
    except Exception as e:
        print(f"    Batch error: {e}")
        return {}

    results = {}
    if len(tickers) == 1:
        ticker = tickers[0]
        if not data.empty and len(data) > 50:
            results[ticker] = data
    else:
        for ticker in tickers:
            try:
                df = data[ticker].dropna(how="all")
                if not df.empty and len(df) > 50:
                    results[ticker] = df
            except (KeyError, Exception):
                pass
    return results


def download_all(tickers, signal_date):
    start = (signal_date - timedelta(days=HISTORY_YEARS * 365)).strftime("%Y-%m-%d")
    end = datetime.now().strftime("%Y-%m-%d")

    batches = [tickers[i:i + BATCH_SIZE] for i in range(0, len(tickers), BATCH_SIZE)]
    all_data = {}
    failed = []

    print(f"\n  Downloading {len(tickers)} stocks in {len(batches)} batches...")
    print(f"  Date range: {start} → {end}")
    print(f"  Random delays (2-6s) between batches.\n")

    for idx, batch in enumerate(batches):
        pct = (idx + 1) / len(batches) * 100
        sys.stdout.write(f"\r  Batch {idx+1}/{len(batches)} ({pct:.0f}%) — "
                         f"{len(all_data)} loaded")
        sys.stdout.flush()

        batch_data = download_batch(batch, start, end)
        if not batch_data:
            failed.append(batch)
        else:
            all_data.update(batch_data)

        if idx < len(batches) - 1:
            time.sleep(random.uniform(2.0, 6.0))

    if failed:
        print(f"\n\n  Retrying {len(failed)} failed batch(es)...")
        for idx, batch in enumerate(failed):
            time.sleep(random.uniform(8.0, 15.0))
            sys.stdout.write(f"\r  Retry {idx+1}/{len(failed)}...")
            sys.stdout.flush()
            batch_data = download_batch(batch, start, end)
            if batch_data:
                all_data.update(batch_data)

    print(f"\n  Downloaded {len(all_data)} stocks with data\n")
    return all_data


# ══════════════════════════════════════════════════════════════
#  HELPER: standardize daily DataFrame
# ══════════════════════════════════════════════════════════════

def prep_daily(daily: pd.DataFrame) -> pd.DataFrame | None:
    """Standardize columns, return cleaned DataFrame or None."""
    if isinstance(daily.columns, pd.MultiIndex):
        daily.columns = daily.columns.get_level_values(-1)
    daily.columns = [c.lower().strip() for c in daily.columns]

    required = {"open", "high", "low", "close", "volume"}
    if not required.issubset(set(daily.columns)):
        return None

    daily = daily.dropna(subset=["close"]).copy()
    if not isinstance(daily.index, pd.DatetimeIndex):
        daily.index = pd.to_datetime(daily.index)
    return daily


# ══════════════════════════════════════════════════════════════
#  ENTRY SCREEN (same logic as backtest.py)
# ══════════════════════════════════════════════════════════════

def check_entry(daily: pd.DataFrame, signal_date: pd.Timestamp) -> dict | None:
    """
    Check if a stock triggers the oversold filter + crossover signal
    as of signal_date. Returns entry info dict or None.
    """
    daily_to_signal = daily[daily.index <= signal_date]
    if len(daily_to_signal) < 400:
        return None

    signal_price = daily_to_signal["close"].iloc[-1]
    avg_vol = daily_to_signal["volume"].tail(20).mean()
    if signal_price < MIN_PRICE or avg_vol < MIN_AVG_VOLUME:
        return None

    qdf = resample_ohlcv(daily_to_signal, 'QE')
    min_bars = RSI_PERIOD + RSI_SIGNAL_PERIOD + 2
    if len(qdf) < min_bars:
        return None

    qdf = compute_all(
        qdf,
        rsi_period=RSI_PERIOD, rsi_signal_period=RSI_SIGNAL_PERIOD,
        stoch_k=STOCH_K_PERIOD, stoch_slow_k=STOCH_SLOW_K,
        stoch_slow_d=STOCH_SLOW_D,
        smi_period=SMI_PERIOD, smi_smooth1=SMI_SMOOTH1,
        smi_smooth2=SMI_SMOOTH2,
    )

    for col in ['RSI', 'RSI_signal', 'K', 'D', 'SMI']:
        if pd.isna(qdf[col].iloc[-1]) or pd.isna(qdf[col].iloc[-2]):
            return None

    curr = qdf.iloc[-1]
    prev = qdf.iloc[-2]

    # FILTER — at least 1 of 3 oversold
    current_smi = curr['SMI']
    current_k = curr['K']
    current_d = curr['D']
    current_stoch_avg = (current_k + current_d) / 2.0
    current_rsi = curr['RSI']
    current_rsi_sig = curr['RSI_signal']

    smi_below = current_smi < SMI_THRESHOLD
    stoch_below = current_stoch_avg < STOCH_AVG_THRESHOLD
    rsi_below = current_rsi < RSI_AVG_THRESHOLD

    if not (smi_below or stoch_below or rsi_below):
        return None

    # SIGNAL — at least 1 crossover
    stoch_cross = (prev['K'] <= prev['D']) and (current_k > current_d)
    rsi_cross = (prev['RSI'] <= prev['RSI_signal']) and (current_rsi > current_rsi_sig)

    if not (stoch_cross or rsi_cross):
        return None

    signals = []
    if stoch_cross:
        signals.append("Stoch %K > %D")
    if rsi_cross:
        signals.append("RSI > Signal")

    entry_price = curr['close']
    if entry_price <= 0 or pd.isna(entry_price):
        return None

    return {
        "entry_date": qdf.index[-1],
        "entry_price": entry_price,
        "smi": current_smi,
        "stoch_avg": current_stoch_avg,
        "rsi": current_rsi,
        "signal": ", ".join(signals),
    }


# ══════════════════════════════════════════════════════════════
#  OVERBOUGHT HELPERS
# ══════════════════════════════════════════════════════════════

def count_overbought(smi, stoch_avg, rsi) -> int:
    """Count how many of 3 indicators are above overbought."""
    count = 0
    if smi > SMI_OVERBOUGHT:
        count += 1
    if stoch_avg > STOCH_OVERBOUGHT:
        count += 1
    if rsi > RSI_OVERBOUGHT:
        count += 1
    return count


def find_overbought_then_fall(bars_df: pd.DataFrame,
                              entry_date: pd.Timestamp,
                              fell_below_enabled: bool = True,
                              fell_below_bars: int = 1) -> dict | None:
    """
    Walk bars after entry_date. Two possible exit triggers:

    Path A — Fell Below Average (early exit, if enabled):
      Before ever reaching overbought (2 of 3 above thresholds), if
      Stochastic %K falls below %D  AND  RSI falls below RSI_signal
      for `fell_below_bars` consecutive bars, sell.
      Status = "Fell Below Average - Sold"

    Path B — Overbought Reversal:
      Phase 1: wait for 2 of 3 indicators to go ABOVE overbought
      Phase 2: once overbought, sell at the close of the first bar
               where fewer than 2 of 3 remain above overbought.
               Status = "Overbought - Sold"

    Once the stock enters overbought territory, the "fell below average"
    check is disabled — you only sell on the overbought reversal.

    Returns dict with exit info + "status" key, or None (still holding).
    """
    post_entry = bars_df[bars_df.index > entry_date]
    was_overbought = False
    consecutive_below = 0  # tracks consecutive bars below average

    for idx in post_entry.index:
        row = bars_df.loc[idx]
        if pd.isna(row['SMI']) or pd.isna(row['K']) or pd.isna(row['RSI']):
            continue

        smi = row['SMI']
        k = row['K']
        d = row['D']
        stoch_avg = (k + d) / 2.0
        rsi = row['RSI']
        rsi_signal = row['RSI_signal']
        ob_count = count_overbought(smi, stoch_avg, rsi)

        if not was_overbought:
            # Check if we've entered overbought territory
            if ob_count >= 2:
                was_overbought = True
                consecutive_below = 0  # reset — no longer relevant
                # Don't sell this bar — we just got overbought,
                # wait for the reversal on a subsequent bar
            elif fell_below_enabled:
                # Not yet overbought — check for fell-below-average
                # Both must be true: Stoch %K below %D AND RSI below RSI_signal
                stoch_below_avg = k < d
                rsi_below_avg = rsi < rsi_signal if not pd.isna(rsi_signal) else False

                if stoch_below_avg and rsi_below_avg:
                    consecutive_below += 1
                    if consecutive_below >= fell_below_bars:
                        return {
                            "exit_date": idx,
                            "exit_price": row['close'],
                            "exit_smi": smi,
                            "exit_stoch_avg": stoch_avg,
                            "exit_rsi": rsi,
                            "status": "Fell Below Average - Sold",
                        }
                else:
                    consecutive_below = 0  # reset streak
        else:
            # Was overbought — sell when it falls back below
            if ob_count < 2:
                return {
                    "exit_date": idx,
                    "exit_price": row['close'],
                    "exit_smi": smi,
                    "exit_stoch_avg": stoch_avg,
                    "exit_rsi": rsi,
                    "status": "Overbought - Sold",
                }

    return None  # never triggered either exit


# ══════════════════════════════════════════════════════════════
#  EXIT SCENARIO: QUARTERLY OVERBOUGHT → FALL BELOW
# ══════════════════════════════════════════════════════════════

def find_quarterly_exit(daily: pd.DataFrame, entry_date: pd.Timestamp,
                        fell_below_enabled: bool = True,
                        fell_below_bars: int = 1) -> dict | None:
    """
    On quarterly bars after entry: two exit paths.
    fell_below_bars = number of consecutive quarters below average to trigger.
    """
    qdf = resample_ohlcv(daily, 'QE')
    min_bars = RSI_PERIOD + RSI_SIGNAL_PERIOD + 2
    if len(qdf) < min_bars:
        return None

    qdf = compute_all(
        qdf,
        rsi_period=RSI_PERIOD, rsi_signal_period=RSI_SIGNAL_PERIOD,
        stoch_k=STOCH_K_PERIOD, stoch_slow_k=STOCH_SLOW_K,
        stoch_slow_d=STOCH_SLOW_D,
        smi_period=SMI_PERIOD, smi_smooth1=SMI_SMOOTH1,
        smi_smooth2=SMI_SMOOTH2,
    )

    return find_overbought_then_fall(qdf, entry_date,
                                     fell_below_enabled, fell_below_bars)


# ══════════════════════════════════════════════════════════════
#  EXIT SCENARIO: MONTHLY OVERBOUGHT → FALL BELOW
# ══════════════════════════════════════════════════════════════

def find_monthly_exit(daily: pd.DataFrame, entry_date: pd.Timestamp,
                      fell_below_enabled: bool = True,
                      fell_below_bars: int = 1) -> dict | None:
    """
    On monthly bars after entry: two exit paths.
    fell_below_bars = number of consecutive months below average to trigger.
    """
    mdf = resample_ohlcv(daily, 'ME')
    min_bars = RSI_PERIOD + RSI_SIGNAL_PERIOD + 2
    if len(mdf) < min_bars:
        return None

    mdf = compute_all(
        mdf,
        rsi_period=RSI_PERIOD, rsi_signal_period=RSI_SIGNAL_PERIOD,
        stoch_k=STOCH_K_PERIOD, stoch_slow_k=STOCH_SLOW_K,
        stoch_slow_d=STOCH_SLOW_D,
        smi_period=SMI_PERIOD, smi_smooth1=SMI_SMOOTH1,
        smi_smooth2=SMI_SMOOTH2,
    )

    return find_overbought_then_fall(mdf, entry_date,
                                     fell_below_enabled, fell_below_bars)


# ══════════════════════════════════════════════════════════════
#  COMPUTE RETURN
# ══════════════════════════════════════════════════════════════

def calc_return(entry_price, exit_price, entry_date, exit_date):
    """Compute total and annualized return."""
    if entry_price <= 0 or pd.isna(entry_price):
        return 0.0, 0.0, 0.0

    total_pct = ((exit_price - entry_price) / entry_price) * 100
    days_held = (exit_date - entry_date).days
    years_held = max(days_held / 365.25, 0.01)

    if total_pct > -100:
        annualized = ((1 + total_pct / 100) ** (1 / years_held) - 1) * 100
    else:
        annualized = -100.0

    return total_pct, annualized, years_held


# ══════════════════════════════════════════════════════════════
#  GENERATE ALL QUARTER-END DATES FROM START TO TODAY
# ══════════════════════════════════════════════════════════════

def get_quarter_ends(start_date: pd.Timestamp) -> list[pd.Timestamp]:
    """
    Return all quarter-end dates from start_date through the last
    COMPLETED quarter (at least 1 quarter before today).
    """
    today = pd.Timestamp.now().normalize()
    # Last completed quarter end
    current_qe = today - pd.offsets.QuarterEnd(1)
    if current_qe >= today:
        current_qe = today - pd.offsets.QuarterEnd(2)

    # Generate range
    dates = pd.date_range(start=start_date, end=current_qe, freq='QE')
    return list(dates)


# ══════════════════════════════════════════════════════════════
#  PROCESS ONE STOCK — ROLLING ENTRIES ACROSS ALL QUARTERS
# ══════════════════════════════════════════════════════════════

def process_stock(ticker: str, daily: pd.DataFrame,
                  signal_date: pd.Timestamp,
                  fell_below_enabled: bool = True,
                  fell_below_bars: int = 1) -> list[dict]:
    """
    Check every completed quarter from signal_date through today.
    Any quarter that triggers entry → compute all 3 exit scenarios.
    Returns a list of result dicts (possibly multiple entries per stock).
    """
    daily = prep_daily(daily)
    if daily is None:
        return []

    quarter_ends = get_quarter_ends(signal_date)
    if not quarter_ends:
        return []

    today_price = daily["close"].iloc[-1]
    today_date = daily.index[-1]

    results = []

    for qe in quarter_ends:
        entry = check_entry(daily, qe)
        if entry is None:
            continue

        entry_date = entry["entry_date"]
        entry_price = entry["entry_price"]

        # ── Scenario 1: Hold Through Today ──
        hold_total, hold_ann, hold_years = calc_return(
            entry_price, today_price, entry_date, today_date)

        # ── Scenario 2: Quarterly Overbought Exit ──
        q_exit = find_quarterly_exit(daily, entry_date,
                                      fell_below_enabled, fell_below_bars)
        if q_exit:
            q_total, q_ann, q_years = calc_return(
                entry_price, q_exit["exit_price"], entry_date, q_exit["exit_date"])
            q_exit_date = q_exit["exit_date"]
            q_exit_price = q_exit["exit_price"]
            q_exit_smi = q_exit["exit_smi"]
            q_exit_stoch = q_exit["exit_stoch_avg"]
            q_exit_rsi = q_exit["exit_rsi"]
            q_status = q_exit["status"]
        else:
            q_total, q_ann, q_years = hold_total, hold_ann, hold_years
            q_exit_date = today_date
            q_exit_price = today_price
            q_exit_smi = np.nan
            q_exit_stoch = np.nan
            q_exit_rsi = np.nan
            q_status = "Still Holding"

        # ── Scenario 3: Monthly Overbought Exit ──
        m_exit = find_monthly_exit(daily, entry_date,
                                    fell_below_enabled, fell_below_bars)
        if m_exit:
            m_total, m_ann, m_years = calc_return(
                entry_price, m_exit["exit_price"], entry_date, m_exit["exit_date"])
            m_exit_date = m_exit["exit_date"]
            m_exit_price = m_exit["exit_price"]
            m_exit_smi = m_exit["exit_smi"]
            m_exit_stoch = m_exit["exit_stoch_avg"]
            m_exit_rsi = m_exit["exit_rsi"]
            m_status = m_exit["status"]
        else:
            m_total, m_ann, m_years = hold_total, hold_ann, hold_years
            m_exit_date = today_date
            m_exit_price = today_price
            m_exit_smi = np.nan
            m_exit_stoch = np.nan
            m_exit_rsi = np.nan
            m_status = "Still Holding"

        results.append({
            "Ticker": ticker,
            "Entry Date": entry_date,
            "Entry Price": round(entry_price, 2),
            "Entry SMI": round(entry["smi"], 2),
            "Entry Stoch Avg": round(entry["stoch_avg"], 2),
            "Entry RSI": round(entry["rsi"], 2),
            "Entry Signal": entry["signal"],
            # Hold through today
            "Hold Current Price": round(today_price, 2),
            "Hold Total Return %": round(hold_total, 1),
            "Hold Annualized %": round(hold_ann, 1),
            "Hold Years": round(hold_years, 1),
            # Quarterly exit
            "Q Status": q_status,
            "Q Exit Date": q_exit_date,
            "Q Exit Price": round(q_exit_price, 2),
            "Q Exit SMI": round(q_exit_smi, 2) if not pd.isna(q_exit_smi) else "",
            "Q Exit Stoch Avg": round(q_exit_stoch, 2) if not pd.isna(q_exit_stoch) else "",
            "Q Exit RSI": round(q_exit_rsi, 2) if not pd.isna(q_exit_rsi) else "",
            "Q Total Return %": round(q_total, 1),
            "Q Annualized %": round(q_ann, 1),
            "Q Years Held": round(q_years, 1),
            # Monthly exit
            "M Status": m_status,
            "M Exit Date": m_exit_date,
            "M Exit Price": round(m_exit_price, 2),
            "M Exit SMI": round(m_exit_smi, 2) if not pd.isna(m_exit_smi) else "",
            "M Exit Stoch Avg": round(m_exit_stoch, 2) if not pd.isna(m_exit_stoch) else "",
            "M Exit RSI": round(m_exit_rsi, 2) if not pd.isna(m_exit_rsi) else "",
            "M Total Return %": round(m_total, 1),
            "M Annualized %": round(m_ann, 1),
            "M Years Held": round(m_years, 1),
        })

    return results


# ══════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════

def write_excel(results_df: pd.DataFrame, signal_date_str: str,
                exchanges: list[str], out_file: str,
                fell_below_enabled: bool = True, fell_below_bars: int = 1):
    """Write the 4-tab Excel workbook."""

    wb = Workbook()

    # ── Styles ──
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(horizontal="center", vertical="center")
    pct_fmt = '0.0"%"'
    price_fmt = '$#,##0.00'
    date_fmt = 'YYYY-MM-DD'
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # Green/red conditional fills
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    sold_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    holding_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    title_font = Font(name="Arial", bold=True, size=14, color="2F5496")
    subtitle_font = Font(name="Arial", size=11, color="595959")
    metric_label_font = Font(name="Arial", bold=True, size=11)
    metric_val_font = Font(name="Arial", size=11)

    def auto_width(ws, min_width=10, max_width=22):
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = min_width
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
            ws.column_dimensions[col_letter].width = max_len

    def write_data_sheet(ws, df, columns, col_formats, sheet_title,
                         status_col=None):
        """Write a data tab with header row and formatted data."""
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=sheet_title)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="left")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        sub_cell = ws.cell(row=2, column=1,
                           value=f"Starting from: {signal_date_str}  |  "
                                 f"Exchange(s): {', '.join(exchanges)}  |  "
                                 f"Trades: {len(df)}")
        sub_cell.font = subtitle_font

        header_row = 4
        data_start = 5

        # Headers
        for ci, col_name in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=ci, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data
        for ri, (_, row) in enumerate(df.iterrows(), data_start):
            for ci, col_name in enumerate(columns, 1):
                val = row.get(col_name, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

                # Apply format
                fmt = col_formats.get(col_name)
                if fmt:
                    cell.number_format = fmt

                # Color return columns
                if "Return %" in col_name and isinstance(val, (int, float)):
                    cell.fill = green_fill if val > 0 else red_fill

                # Color status column
                if status_col and col_name == status_col:
                    if "Overbought" in str(val):
                        cell.fill = sold_fill
                    elif "Fell Below" in str(val):
                        cell.fill = PatternFill(start_color="FFE0B2",
                                                end_color="FFE0B2",
                                                fill_type="solid")  # orange tint
                    elif val == "Still Holding":
                        cell.fill = holding_fill

        auto_width(ws)

    # ── Sort by hold return (descending) for consistent ordering ──
    results_df = results_df.sort_values("Hold Total Return %",
                                        ascending=False).reset_index(drop=True)

    # ════════════════════════════════════════
    #  TAB 1: SUMMARY
    # ════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.merge_cells("A1:G1")
    ws_sum.cell(row=1, column=1,
                value="Backtest Scenarios — Summary").font = title_font

    ws_sum.merge_cells("A2:G2")
    unique_tickers = results_df["Ticker"].nunique()
    ws_sum.cell(row=2, column=1,
                value=f"Starting from: {signal_date_str}  |  "
                      f"Exchange(s): {', '.join(exchanges)}  |  "
                      f"Total trades: {len(results_df)}  |  "
                      f"Unique tickers: {unique_tickers}").font = subtitle_font

    ws_sum.merge_cells("A3:G3")
    ws_sum.cell(row=3, column=1,
                value=f"Overbought exit: 2 of 3 indicators above "
                      f"SMI>{SMI_OVERBOUGHT}, Stoch Avg>{STOCH_OVERBOUGHT}, "
                      f"RSI>{RSI_OVERBOUGHT}").font = subtitle_font

    fell_below_label = "OFF"
    if fell_below_enabled:
        fell_below_label = (f"ON — sell after {fell_below_bars} consecutive "
                            f"bar(s) with Stoch %K < %D AND RSI < Signal")
    ws_sum.merge_cells("A4:G4")
    ws_sum.cell(row=4, column=1,
                value=f"Fell-below-average early exit: "
                      f"{fell_below_label}").font = subtitle_font

    # Summary table
    summary_headers = [
        "Metric",
        "Hold Through Today",
        "Quarterly Overbought Exit",
        "Monthly Overbought Exit",
    ]
    header_row = 6
    for ci, h in enumerate(summary_headers, 1):
        cell = ws_sum.cell(row=header_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Compute stats
    def scenario_stats(total_col, ann_col, status_col=None):
        totals = results_df[total_col]
        anns = results_df[ann_col]
        stats = {
            "Total Stocks": len(results_df),
            "Average Total Return %": round(totals.mean(), 1),
            "Median Total Return %": round(totals.median(), 1),
            "Avg Annualized Return %": round(anns.mean(), 1),
            "% Positive": round((totals > 0).sum() / len(totals) * 100, 0),
            "Best Return %": round(totals.max(), 1),
            "Worst Return %": round(totals.min(), 1),
        }
        if status_col:
            ob_sold = results_df[status_col].str.contains("Overbought", na=False).sum()
            avg_sold = results_df[status_col].str.contains("Fell Below", na=False).sum()
            still = (results_df[status_col] == "Still Holding").sum()
            stats["Overbought - Sold"] = ob_sold
            stats["Fell Below Average - Sold"] = avg_sold
            stats["Still Holding"] = still
            # Avg return by exit type
            ob_mask = results_df[status_col].str.contains("Overbought", na=False)
            avg_mask = results_df[status_col].str.contains("Fell Below", na=False)
            still_mask = results_df[status_col] == "Still Holding"
            if ob_mask.sum() > 0:
                stats["Avg Return (Overbought Sold) %"] = round(
                    totals[ob_mask].mean(), 1)
            else:
                stats["Avg Return (Overbought Sold) %"] = "N/A"
            if avg_mask.sum() > 0:
                stats["Avg Return (Fell Below Avg Sold) %"] = round(
                    totals[avg_mask].mean(), 1)
            else:
                stats["Avg Return (Fell Below Avg Sold) %"] = "N/A"
            if still_mask.sum() > 0:
                stats["Avg Return (Still Holding) %"] = round(
                    totals[still_mask].mean(), 1)
            else:
                stats["Avg Return (Still Holding) %"] = "N/A"
        return stats

    hold_stats = scenario_stats("Hold Total Return %", "Hold Annualized %")
    q_stats = scenario_stats("Q Total Return %", "Q Annualized %", "Q Status")
    m_stats = scenario_stats("M Total Return %", "M Annualized %", "M Status")

    # Build all metric rows (union of all keys, ordered)
    all_metrics = list(hold_stats.keys())
    for k in q_stats:
        if k not in all_metrics:
            all_metrics.append(k)
    for k in m_stats:
        if k not in all_metrics:
            all_metrics.append(k)

    for ri, metric in enumerate(all_metrics, header_row + 1):
        ws_sum.cell(row=ri, column=1, value=metric).font = metric_label_font
        ws_sum.cell(row=ri, column=1).border = thin_border

        for ci, stats in enumerate([hold_stats, q_stats, m_stats], 2):
            val = stats.get(metric, "")
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            cell.font = metric_val_font
            cell.alignment = data_align
            cell.border = thin_border
            if isinstance(val, (int, float)) and "Return" in metric:
                cell.fill = green_fill if val > 0 else red_fill

    # Distribution table
    dist_row = header_row + len(all_metrics) + 3
    ws_sum.merge_cells(start_row=dist_row, start_column=1,
                       end_row=dist_row, end_column=4)
    ws_sum.cell(row=dist_row, column=1,
                value="Return Distribution").font = Font(
                    name="Arial", bold=True, size=12, color="2F5496")

    dist_headers = ["Range", "Hold Today", "Q Exit", "M Exit"]
    for ci, h in enumerate(dist_headers, 1):
        cell = ws_sum.cell(row=dist_row + 1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    bins = [
        ("> +500%",    lambda s: s > 500),
        ("+200–500%",  lambda s: (s > 200) & (s <= 500)),
        ("+100–200%",  lambda s: (s > 100) & (s <= 200)),
        ("+50–100%",   lambda s: (s > 50) & (s <= 100)),
        ("+0–50%",     lambda s: (s > 0) & (s <= 50)),
        ("-0–50%",     lambda s: (s <= 0) & (s > -50)),
        ("< -50%",     lambda s: s <= -50),
    ]

    for bi, (label, cond) in enumerate(bins, dist_row + 2):
        ws_sum.cell(row=bi, column=1, value=label).font = data_font
        ws_sum.cell(row=bi, column=1).border = thin_border

        for ci, col_name in enumerate(["Hold Total Return %",
                                        "Q Total Return %",
                                        "M Total Return %"], 2):
            count = int(cond(results_df[col_name]).sum())
            cell = ws_sum.cell(row=bi, column=ci, value=count)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    auto_width(ws_sum, min_width=14, max_width=30)

    # ════════════════════════════════════════
    #  TAB 2: HOLD THROUGH TODAY
    # ════════════════════════════════════════
    ws_hold = wb.create_sheet("Hold Through Today")
    hold_cols = [
        "Ticker", "Entry Date", "Entry Price", "Entry SMI", "Entry Stoch Avg",
        "Entry RSI", "Entry Signal", "Hold Current Price",
        "Hold Total Return %", "Hold Annualized %", "Hold Years",
    ]
    hold_fmts = {
        "Entry Price": price_fmt, "Hold Current Price": price_fmt,
        "Entry Date": date_fmt,
    }
    hold_sorted = results_df.sort_values("Hold Total Return %",
                                          ascending=False)
    write_data_sheet(ws_hold, hold_sorted, hold_cols, hold_fmts,
                     "Scenario 1: Hold Through Today")

    # ════════════════════════════════════════
    #  TAB 3: QUARTERLY OVERBOUGHT EXIT
    # ════════════════════════════════════════
    ws_q = wb.create_sheet("Sell Quarterly Overbought")
    q_cols = [
        "Ticker", "Entry Date", "Entry Price", "Entry Signal",
        "Q Status", "Q Exit Date", "Q Exit Price",
        "Q Exit SMI", "Q Exit Stoch Avg", "Q Exit RSI",
        "Q Total Return %", "Q Annualized %", "Q Years Held",
    ]
    q_fmts = {
        "Entry Price": price_fmt, "Q Exit Price": price_fmt,
        "Entry Date": date_fmt, "Q Exit Date": date_fmt,
    }
    q_sorted = results_df.sort_values("Q Total Return %", ascending=False)
    write_data_sheet(ws_q, q_sorted, q_cols, q_fmts,
                     "Scenario 2: Sell on Quarterly Overbought "
                     f"(2 of 3: SMI>{SMI_OVERBOUGHT}, "
                     f"Stoch>{STOCH_OVERBOUGHT}, RSI>{RSI_OVERBOUGHT})",
                     status_col="Q Status")

    # ════════════════════════════════════════
    #  TAB 4: MONTHLY OVERBOUGHT EXIT
    # ════════════════════════════════════════
    ws_m = wb.create_sheet("Sell Monthly Overbought")
    m_cols = [
        "Ticker", "Entry Date", "Entry Price", "Entry Signal",
        "M Status", "M Exit Date", "M Exit Price",
        "M Exit SMI", "M Exit Stoch Avg", "M Exit RSI",
        "M Total Return %", "M Annualized %", "M Years Held",
    ]
    m_fmts = {
        "Entry Price": price_fmt, "M Exit Price": price_fmt,
        "Entry Date": date_fmt, "M Exit Date": date_fmt,
    }
    m_sorted = results_df.sort_values("M Total Return %", ascending=False)
    write_data_sheet(ws_m, m_sorted, m_cols, m_fmts,
                     "Scenario 3: Sell on Monthly Overbought "
                     f"(2 of 3: SMI>{SMI_OVERBOUGHT}, "
                     f"Stoch>{STOCH_OVERBOUGHT}, RSI>{RSI_OVERBOUGHT})",
                     status_col="M Status")

    # ── Save ──
    wb.save(out_file)


# ══════════════════════════════════════════════════════════════
#  CONSOLE SUMMARY
# ══════════════════════════════════════════════════════════════

def print_summary(results_df, signal_date_str):
    total = len(results_df)
    if total == 0:
        print("\n  No stocks triggered the screener as of that date.")
        return

    print(f"\n{'=' * 78}")
    print(f"  BACKTEST SCENARIOS — Entries from {signal_date_str} through today")
    print(f"{'=' * 78}")
    unique_tickers = results_df["Ticker"].nunique()
    print(f"  Total trades (entries): {total}  "
          f"({unique_tickers} unique tickers)")
    print()

    for label, tcol, acol, scol in [
        ("Hold Through Today",        "Hold Total Return %", "Hold Annualized %", None),
        ("Quarterly Overbought Exit", "Q Total Return %",    "Q Annualized %",    "Q Status"),
        ("Monthly Overbought Exit",   "M Total Return %",    "M Annualized %",    "M Status"),
    ]:
        avg_ret = results_df[tcol].mean()
        med_ret = results_df[tcol].median()
        pct_pos = (results_df[tcol] > 0).sum() / total * 100
        avg_ann = results_df[acol].mean()

        print(f"  ── {label} ──")
        print(f"     Avg total return:  {avg_ret:>+.1f}%")
        print(f"     Median return:     {med_ret:>+.1f}%")
        print(f"     Avg annualized:    {avg_ann:>+.1f}%")
        print(f"     % positive:        {pct_pos:.0f}%")

        if scol:
            ob_sold = results_df[scol].str.contains("Overbought", na=False).sum()
            avg_sold = results_df[scol].str.contains("Fell Below", na=False).sum()
            holding = (results_df[scol] == "Still Holding").sum()
            print(f"     Overbought sold: {ob_sold}  |  "
                  f"Fell below avg sold: {avg_sold}  |  "
                  f"Still holding: {holding}")
        print()


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description='Backtest: oversold entry → overbought exit scenarios')
    ap.add_argument('--date', required=True,
                    help='Signal date: "Q1 2015", "Q4 2008", "March 2020", "2015-06-30"')
    ap.add_argument('--out', default=None,
                    help='Output Excel filename (default: auto-generated)')
    ap.add_argument('--fell-below', dest='fell_below', action='store_true', default=True,
                    help='Enable fell-below-average early exit (default: enabled)')
    ap.add_argument('--no-fell-below', dest='fell_below', action='store_false',
                    help='Disable fell-below-average early exit (overbought reversal only)')
    ap.add_argument('--fell-below-bars', type=int, default=1,
                    help='Consecutive bars below average before selling '
                         '(quarters for quarterly, months for monthly; default: 1)')
    args = ap.parse_args()

    signal_date = parse_date(args.date)
    signal_ts = pd.Timestamp(signal_date)

    fell_below_label = "OFF"
    if args.fell_below:
        fell_below_label = f"ON ({args.fell_below_bars} consecutive bar(s))"

    print()
    print("=" * 78)
    print("  BACKTEST SCENARIOS — Oversold Entry → Overbought Exit")
    print(f"  Starting from:  {signal_date.strftime('%Y-%m-%d')} → today")
    print(f"  Scans every completed quarter for new entries")
    print(f"  Overbought thresholds (2 of 3): SMI > {SMI_OVERBOUGHT}, "
          f"Stoch Avg > {STOCH_OVERBOUGHT}, RSI > {RSI_OVERBOUGHT}")
    print(f"  Fell-below-average early exit: {fell_below_label}")
    print("=" * 78)
    print()
    print("  ⚠️  SURVIVORSHIP BIAS WARNING: This only tests stocks still listed today.")
    print("     Stocks that went bankrupt or were delisted after the signal date are")
    print("     missing, which makes results look better than they actually were.")
    print()

    # ── Exchange selection ──
    exchanges = prompt_exchange()

    # ── Get tickers ──
    print()
    tickers = get_stock_tickers(exchanges)

    # ── Download ──
    all_data = download_all(tickers, signal_date)

    # ── Screen + compute scenarios ──
    quarter_ends = get_quarter_ends(signal_ts)
    print(f"  Scanning {len(quarter_ends)} quarters from "
          f"{quarter_ends[0].strftime('%Y-%m-%d')} to "
          f"{quarter_ends[-1].strftime('%Y-%m-%d')}")
    print(f"  Screening + computing exit scenarios...")
    results = []
    for i, (ticker, df) in enumerate(all_data.items()):
        if (i + 1) % 200 == 0:
            sys.stdout.write(f"\r  Processed {i+1}/{len(all_data)} stocks — "
                             f"{len(results)} trades found...")
            sys.stdout.flush()

        stock_results = process_stock(ticker, df, signal_ts,
                                      fell_below_enabled=args.fell_below,
                                      fell_below_bars=args.fell_below_bars)
        results.extend(stock_results)

    print(f"\r  Processed {len(all_data)}/{len(all_data)} stocks — "
          f"{len(results)} trades found")

    if not results:
        print("\n  No stocks triggered the screener across any quarter in that range.")
        print("  Try a different starting date or a wider range.")
        return

    results_df = pd.DataFrame(results)

    # ── Console summary ──
    print_summary(results_df, signal_date.strftime('%Y-%m-%d'))

    # ── Excel output ──
    if args.out:
        out_file = args.out
    else:
        date_tag = signal_date.strftime('%Y%m%d')
        exch_tag = "_".join(e.lower() for e in exchanges)
        out_file = f"backtest_scenarios_{exch_tag}_{date_tag}.xlsx"

    write_excel(results_df, signal_date.strftime('%Y-%m-%d'), exchanges, out_file,
                fell_below_enabled=args.fell_below,
                fell_below_bars=args.fell_below_bars)
    unique_tickers = results_df["Ticker"].nunique()
    print(f"  Excel saved → {out_file}")
    print(f"  ({len(results_df)} trades across {unique_tickers} unique tickers, 4 tabs)")
    print()


if __name__ == "__main__":
    main()